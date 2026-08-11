import streamlit as st
import pandas as pd
import difflib
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="نظام تحليل مخازن المقاولات", layout="wide")
st.markdown("<style>body, .stApp { direction: rtl; text-align: right; }</style>", unsafe_allow_html=True)

DEAD_STOCK_DAYS = 60

# ---------------- محرك التنظيف ----------------
def find_duplicate_item_names(item_names, threshold=0.6):
    unique_names = list(set(item_names))
    checked, groups = set(), []
    for name in unique_names:
        if name in checked:
            continue
        similar = [name]
        checked.add(name)
        for other in unique_names:
            if other in checked:
                continue
            if difflib.SequenceMatcher(None, name, other).ratio() >= threshold:
                similar.append(other)
                checked.add(other)
        if len(similar) > 1:
            groups.append(similar)
    return groups


def clean_transactions(df, name_mapping=None):
    df = df.copy()
    df["التاريخ"] = pd.to_datetime(df["التاريخ"])
    if name_mapping:
        df["اسم الصنف"] = df["اسم الصنف"].replace(name_mapping)
    errors = df[df["الكمية"] <= 0].copy()
    df = df[df["الكمية"] > 0]
    return df, errors


# ---------------- محرك التحليل ----------------
def compute_stock_balance(df):
    pivot = df.pivot_table(index=["المخزن", "اسم الصنف"], columns="نوع الحركة",
                            values="الكمية", aggfunc="sum", fill_value=0)
    for col in ["استلام", "صرف"]:
        if col not in pivot.columns:
            pivot[col] = 0
    pivot["الرصيد الحالي"] = pivot["استلام"] - pivot["صرف"]
    return pivot.reset_index()


def detect_dead_stock(df, balances, today):
    issues = df[df["نوع الحركة"] == "صرف"]
    last_issue = issues.groupby(["المخزن", "اسم الصنف"])["التاريخ"].max().reset_index()
    last_issue.columns = ["المخزن", "اسم الصنف", "آخر تاريخ صرف"]
    result = balances.merge(last_issue, on=["المخزن", "اسم الصنف"], how="left")
    result["آخر تاريخ صرف"] = result["آخر تاريخ صرف"].fillna(pd.Timestamp("2020-01-01"))
    result["أيام من غير حركة"] = (today - result["آخر تاريخ صرف"]).dt.days
    avg_price = df.groupby("اسم الصنف")["سعر الوحدة"].mean().reset_index()
    avg_price.columns = ["اسم الصنف", "متوسط السعر"]
    result = result.merge(avg_price, on="اسم الصنف", how="left")
    result["القيمة الراكدة"] = result["الرصيد الحالي"] * result["متوسط السعر"]
    dead = result[(result["الرصيد الحالي"] > 0) & (result["أيام من غير حركة"] >= DEAD_STOCK_DAYS)]
    return dead.sort_values("القيمة الراكدة", ascending=False)


def forecast_consumption(df, item, warehouse, today):
    subset = df[(df["اسم الصنف"] == item) & (df["المخزن"] == warehouse) & (df["نوع الحركة"] == "صرف")]
    if subset.empty:
        return 0.0
    subset = subset.set_index("التاريخ").resample("D")["الكمية"].sum().fillna(0)
    last_90 = subset[subset.index > today - pd.Timedelta(days=90)]
    last_30 = subset[subset.index > today - pd.Timedelta(days=30)]
    if len(last_90) == 0 or last_90.sum() == 0:
        return 0.0
    rate = 0.7 * (last_30.mean() if len(last_30) else 0) + 0.3 * last_90.mean()
    return round(rate, 2)


def build_reorder_report(df, balances, today, lead_time_days=7):
    rows = []
    for _, r in balances.iterrows():
        if r["الرصيد الحالي"] <= 0:
            continue
        rate = forecast_consumption(df, r["اسم الصنف"], r["المخزن"], today)
        if rate <= 0:
            continue
        days_left = r["الرصيد الحالي"] / rate
        rows.append({"المخزن": r["المخزن"], "اسم الصنف": r["اسم الصنف"],
                      "الرصيد الحالي": r["الرصيد الحالي"], "معدل الاستهلاك اليومي": rate,
                      "أيام متبقية للنفاذ": round(days_left, 1),
                      "يحتاج طلب خلال (يوم)": round(days_left - lead_time_days, 1)})
    return pd.DataFrame(rows).sort_values("أيام متبقية للنفاذ") if rows else pd.DataFrame(rows)


def style_and_export(sheets: dict) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, d in sheets.items():
            d.to_excel(writer, sheet_name=name, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill, cell.font = header_fill, header_font
            cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(14, length + 3)
        ws.sheet_view.rightToLeft = True
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ==================== واجهة المستخدم ====================
st.title("📦 نظام تحليل مخازن المقاولات")
st.caption("ارفع شيت الإكسيل بتاع حركة المخازن، والنظام يطلعلك الراكد والتنبؤ بالنفاذ تلقائيًا")

uploaded = st.file_uploader("ارفع ملف الإكسيل", type=["xlsx"])

# خرائط الأعمدة المدعومة: كل خريطة بتمثل صيغة ملف مختلفة (زي تصدير Dynamics 365)
COLUMN_PROFILES = {
    "الصيغة القياسية (التاريخ، المخزن، اسم الصنف...)": {
        "date": "التاريخ", "warehouse": "المخزن", "item": "اسم الصنف",
        "qty": "الكمية", "move_type": "نوع الحركة", "unit_price": "سعر الوحدة",
    },
    "تصدير Dynamics 365 (Physical date, Warehouse, Description...)": {
        "date": "Physical date", "warehouse": "Warehouse", "item": "Description",
        "qty": "Quantity", "move_type": None, "unit_price": None,
        "cost_amount": "Physical cost amount",
    },
}


def normalize_columns(raw: pd.DataFrame, profile: dict) -> pd.DataFrame:
    """يحول أي صيغة ملف للصيغة الداخلية القياسية اللي المحرك شغال بيها"""
    df = pd.DataFrame()
    df["التاريخ"] = pd.to_datetime(raw[profile["date"]])
    df["المخزن"] = raw[profile["warehouse"]].astype(str)
    df["اسم الصنف"] = raw[profile["item"]].astype(str).str.strip()

    qty_raw = pd.to_numeric(raw[profile["qty"]], errors="coerce").fillna(0)

    if profile.get("move_type"):
        df["نوع الحركة"] = raw[profile["move_type"]]
        df["الكمية"] = qty_raw.abs()
    else:
        # مفيش عمود نوع حركة صريح — نستنتجه من إشارة الكمية
        # موجب = استلام (دخول للمخزن) / سالب = صرف (خروج من المخزن)
        df["نوع الحركة"] = qty_raw.apply(lambda x: "استلام" if x >= 0 else "صرف")
        df["الكمية"] = qty_raw.abs()

    if profile.get("unit_price"):
        df["سعر الوحدة"] = pd.to_numeric(raw[profile["unit_price"]], errors="coerce").fillna(0)
    elif profile.get("cost_amount"):
        cost = pd.to_numeric(raw[profile["cost_amount"]], errors="coerce").fillna(0).abs()
        df["سعر الوحدة"] = (cost / df["الكمية"].replace(0, pd.NA)).fillna(0)
    else:
        df["سعر الوحدة"] = 0

    return df[df["الكمية"] > 0].reset_index(drop=True)


if uploaded:
    raw_original = pd.read_excel(uploaded)
    st.subheader("0️⃣ اختار صيغة الأعمدة")
    profile_name = st.selectbox("شكل الأعمدة في ملفك إيه؟", options=list(COLUMN_PROFILES.keys()), index=1)
    profile = COLUMN_PROFILES[profile_name]

    missing = [v for k, v in profile.items() if v and k != "cost_amount" and v not in raw_original.columns]
    # تحقق أدق: بس من الأعمدة الأساسية المطلوبة فعليًا لهذا البروفايل
    required_keys = ["date", "warehouse", "item", "qty"]
    missing = [profile[k] for k in required_keys if profile[k] not in raw_original.columns]
    if missing:
        st.error(f"الملف ناقصه الأعمدة دي: {missing}")
        st.stop()

    raw = normalize_columns(raw_original, profile)
    st.success(f"تم تجهيز {len(raw)} حركة بنجاح من أصل {len(raw_original)} صف")
    with st.expander("👀 شوف عينة من البيانات بعد التحويل"):
        st.dataframe(raw.head(10), use_container_width=True)

    # مراجعة الأصناف المتشابهة
    st.subheader("1️⃣ مراجعة الأصناف المحتمل تكررها")
    groups = find_duplicate_item_names(raw["اسم الصنف"].unique().tolist())
    name_mapping = {}
    if groups:
        st.warning(f"لقينا {len(groups)} مجموعة أصناف متشابهة الاسم — اختار الاسم الصحيح لكل مجموعة")
        for i, g in enumerate(groups):
            chosen = st.selectbox(f"مجموعة {i+1}: {g}", options=g, key=f"grp_{i}")
            for name in g:
                if name != chosen:
                    name_mapping[name] = chosen
    else:
        st.info("مفيش أصناف متشابهة الاسم")

    today = st.date_input("تاريخ التحليل (افتراضيًا اليوم)", value=pd.Timestamp.today())
    today = pd.Timestamp(today)

    if st.button("🔍 شغّل التحليل"):
        clean_df, errors_df = clean_transactions(raw, name_mapping)
        balances = compute_stock_balance(clean_df)
        dead_stock = detect_dead_stock(clean_df, balances, today)
        reorder = build_reorder_report(clean_df, balances, today)

        st.subheader("2️⃣ النتائج")
        total_dead = dead_stock["القيمة الراكدة"].sum() if not dead_stock.empty else 0
        col1, col2, col3 = st.columns(3)
        col1.metric("💰 إجمالي قيمة الراكد", f"{total_dead:,.0f} جنيه")
        col2.metric("📉 عدد الأصناف الراكدة", len(dead_stock))
        col3.metric("⚠️ أصناف قريبة من النفاذ", len(reorder[reorder["يحتاج طلب خلال (يوم)"] <= 0]) if not reorder.empty else 0)

        st.markdown("### 🔴 الأصناف الراكدة")
        st.dataframe(dead_stock, use_container_width=True)

        st.markdown("### 🟡 تنبؤ النفاذ واحتياج الطلب")
        st.dataframe(reorder, use_container_width=True)

        if not errors_df.empty:
            st.markdown("### ⚠️ أخطاء بيانات مكتشفة (كميات سالبة/صفر)")
            st.dataframe(errors_df, use_container_width=True)

        sheets = {"الراكد": dead_stock, "تنبؤ النفاذ": reorder, "أرصدة كل الأصناف": balances}
        if not errors_df.empty:
            sheets["أخطاء بيانات"] = errors_df
        excel_bytes = style_and_export(sheets)
        st.download_button("⬇️ تحميل التقرير الكامل Excel", data=excel_bytes,
                            file_name="تقرير_المخازن_الذكي.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.info("ارفع ملف الإكسيل عشان تبدأ التحليل")
